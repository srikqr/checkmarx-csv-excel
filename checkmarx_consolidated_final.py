#!/usr/bin/env python
"""
checkmarx_consolidated_final.py

Generates a consolidated, layman-friendly Excel report from Checkmarx CSV:
- Auto-installs missing dependencies.
- Filters out entries with empty or 'None' Result Severity.
- Groups findings by Vulnerability Type.
- Lists numbered occurrences per vulnerability, with fields comma-separated and blank lines between instances.
- Outputs columns: Vulnerability Type, Occurrences, Severity (includes 'Information' as a severity), Total Findings.
- Creates separate sheets per severity level including 'Information'.
- Applies professional formatting: frozen pane, autofilter, alternating row shading, and severity color coding.
"""

import sys
import os
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

# 1. Auto-install missing dependencies
for pkg in ("pandas", "openpyxl", "xlsxwriter", "numpy"):
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

# 2. Define occurrence detail fields and mapping for friendly labels
DETAIL_FIELDS = [
    "SrcFileName", "Line", "Column", "NodeId", "Name",
    "DestFileName", "DestLine", "DestColumn", "DestNodeId", "DestName"
]

LABEL_MAP = {
    "SrcFileName":  "Source File",
    "Line":         "Line",
    "Column":       "Column",
    "NodeId":       "Node ID",
    "Name":         "Function",
    "DestFileName": "Dest File",
    "DestLine":     "Dest Line",
    "DestColumn":   "Dest Column",
    "DestNodeId":   "Dest Node ID",
    "DestName":     "Dest Function"
}

# 3. Output columns order
OUTPUT_COLS = [
    "Vulnerability Type",
    "Occurrences",
    "Severity",
    "Total Findings"
]

# 4. Severities and their sheet names and colors, including 'Information'
SEVERITIES = ["Critical", "High", "Medium", "Low", "Information"]

SEV_SHEETS = {
    "Critical":    "Critical Issues",
    "High":        "High Issues",
    "Medium":      "Medium Issues",
    "Low":         "Low Issues",
    "Information": "Information Issues"
}

SEV_COLORS = {
    "Critical":    "FFC7CE",  # light red
    "High":        "FFEB9C",  # light yellow
    "Medium":      "C6EFCE",  # light green
    "Low":         "BDD7EE",  # light blue
    "Information": "D9D9D9"   # light gray
}

# 5. Load and clean CSV
def load_and_clean(csv_path):
    df = pd.read_csv(csv_path, dtype=str, keep_default_na=False).fillna("")
    # Filter out empty or 'None' Result Severity rows
    df = df[~df["Result Severity"].str.strip().str.lower().isin(["", "none"])]
    return df

# 6. Format occurrence details and consolidate by Query
def consolidate(df):
    def fmt_occurrence(row):
        parts = []
        for fld in DETAIL_FIELDS:
            val = row.get(fld, "")
            if val:
                parts.append(f"{LABEL_MAP[fld]}={val}")
        return ", ".join(parts)

    df["OccText"] = df.apply(fmt_occurrence, axis=1)

    grouped = df.groupby("Query").agg({
        "OccText": lambda items: "\n\n".join(f"{i+1}) {txt}" for i, txt in enumerate(items)),
        "Result Severity": list
    }).reset_index().rename(columns={
        "Query": "Vulnerability Type",
        "OccText": "Occurrences",
        "Result Severity": "_sevs"
    })

    grouped["Total Findings"] = grouped["_sevs"].apply(len)

    # Determine highest severity present in _sevs (Information lowest priority)
    def pick_severity(sevs):
        for sev in SEVERITIES[:-1]:  # Check all except 'Information'
            if sev in sevs:
                return sev
        if "Information" in sevs:
            return "Information"
        return ""

    grouped["Severity"] = grouped["_sevs"].apply(pick_severity)

    return grouped[OUTPUT_COLS]

# 7. Styling helper
def style_ws(ws, title=None):
    if title:
        ws.title = title
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="CCCCCC")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    severity_col_idx = OUTPUT_COLS.index("Severity") + 1

    for col_idx, col in enumerate(ws.columns, start=1):
        max_width = 0
        col_letter = get_column_letter(col_idx)
        for row_idx, cell in enumerate(col, start=1):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx > 1:
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                # Color cells in Severity column
                if col_idx == severity_col_idx and cell.value in SEV_COLORS:
                    cell.fill = PatternFill("solid", fgColor=SEV_COLORS[cell.value])
            if cell.value:
                max_width = max(max_width, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_width + 5, 80)

# 8. Build and save the Excel workbook
def build_workbook(df, output_path):
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All Findings"
    ws_all.append(OUTPUT_COLS)

    for row_idx, rec in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(rec, start=1):
            ws_all.cell(row=row_idx, column=col_idx, value=val)

    style_ws(ws_all)

    for severity in SEVERITIES:
        filtered_df = df[df["Severity"] == severity]
        if not filtered_df.empty:
            ws = wb.create_sheet()
            ws.append(OUTPUT_COLS)
            for row_idx, rec in enumerate(filtered_df.itertuples(index=False), start=2):
                for col_idx, val in enumerate(rec, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            style_ws(ws, title=SEV_SHEETS[severity])

    wb.save(output_path)
    print(f"✓ Report generated: {output_path}")

# 9. Main entry point
def main():
    if len(sys.argv) != 3:
        print("Usage: python checkmarx_consolidated_final.py <input.csv> <output.xlsx>")
        sys.exit(1)

    inp, outp = sys.argv[1], sys.argv[2]

    if not os.path.isfile(inp):
        print(f"Error: File not found: {inp}")
        sys.exit(1)

    df = load_and_clean(inp)
    df_final = consolidate(df)
    build_workbook(df_final, outp)

if __name__ == "__main__":
    main()
